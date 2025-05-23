from openpyxl import load_workbook, Workbook
import os

EXCEL_FILE = "student_register.xlsx"

def ask_name():
    while True:
        name = input("Name of student: ")
        if len(name) >= 3:
            return name

def main():
    if os.path.exists(EXCEL_FILE):
        workbook = load_workbook(filename=EXCEL_FILE)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active

        header_row = ["Name of student", "Age", "Qualification"]
        sheet.append(header_row)
    
    name = ask_name()
    new_row = [name, "24", "BE"]
    sheet.append(new_row)

    workbook.save(filename=EXCEL_FILE)


if __name__ == "__main__":
    main()
